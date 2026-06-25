from typing import List
from io import BytesIO
from openpyxl import load_workbook
from pydantic import ValidationError
from fastapi import HTTPException, status

from web_app.src.schemas.participant import Participant


_HEADER_HINTS = ("фио", "имя", "код", "code", "name", "информац")

_FIELD_LABELS = {
    "full_name": "ФИО",
    "extra_info": "Доп. информация",
    "code": "Код",
}

_TYPE_MESSAGES = {
    "missing": "обязательное поле не заполнено",
    "string_type": "значение должно быть строкой",
    "string_too_short": "значение слишком короткое",
    "string_too_long": "значение слишком длинное",
}


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _looks_like_header(row) -> bool:
    joined = " ".join(_cell_to_str(c).lower() for c in row[:3])
    return any(hint in joined for hint in _HEADER_HINTS)


def _translate_error(err) -> str:
    loc = err.get("loc", ())
    field = loc[0] if loc else None
    label = _FIELD_LABELS.get(field, str(field) if field else "Поле")
    msg = _TYPE_MESSAGES.get(err.get("type", ""), "некорректное значение")
    return f"{label}: {msg}"


def _format_validation_error(e: ValidationError) -> str:
    return "; ".join(_translate_error(err) for err in e.errors())


def parse_participants_xlsx(file_bytes: bytes) -> List[Participant]:
    """Парсинг xlsx. Колонки: ФИО, доп. информация, код (опционально)."""
    try:
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Не удалось прочитать xlsx файл")

    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Файл пуст")

    start = 1 if _looks_like_header(rows[0]) else 0

    participants: List[Participant] = []
    errors: List[str] = []

    for idx, row in enumerate(rows[start:], start=start + 1):
        full_name = _cell_to_str(row[0]) if len(row) > 0 else ""
        extra_info = _cell_to_str(row[1]) if len(row) > 1 else ""
        code_raw = _cell_to_str(row[2]) if len(row) > 2 else ""

        if not full_name:
            continue

        try:
            participants.append(
                Participant(
                    full_name=full_name,
                    extra_info=extra_info,
                    code=code_raw or None,  # валидатор сам решит: взять или сгенерировать
                )
            )
        except ValidationError as e:
            errors.append(f"Строка {idx}: {_format_validation_error(e)}")

    if errors:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={"message": "Ошибки в файле", "errors": errors},
        )
    if not participants:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="В файле нет участников")

    return participants