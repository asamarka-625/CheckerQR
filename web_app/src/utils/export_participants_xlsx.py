# Внешние зависимости
from io import BytesIO
from typing import Any, Dict, List
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
# Внутренние модули
from shared_services import generate_qr_bytes
from web_app.src.core import cfg


_HEADERS = ["ФИО", "Доп. информация", "Код", "QR-код"]
_QR_PX = 110  # размер картинки QR в пикселях


def build_participants_xlsx(
    event_title: str,
    participants: List[Dict[str, Any]]
) -> bytes:
    """Сформировать xlsx с участниками: ФИО, доп. инфо, код, QR-код."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Участники"

    # 1. заголовок
    for col, name in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 2. ширины колонок
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18

    # держим ссылки на буферы, чтобы openpyxl успел прочитать их при save()
    buffers: List[BytesIO] = []

    # 3. строки
    for idx, p in enumerate(participants, start=2):
        ws.cell(row=idx, column=1, value=p.get("full_name", "")).alignment = \
            Alignment(vertical="center", wrap_text=True)
        ws.cell(row=idx, column=2, value=p.get("extra_info", "")).alignment = \
            Alignment(vertical="center", wrap_text=True)
        ws.cell(row=idx, column=3, value=p.get("code", "")).alignment = \
            Alignment(horizontal="center", vertical="center")

        # высота строки под QR (px -> points)
        ws.row_dimensions[idx].height = _QR_PX * 0.75

        qr_link = (
            f"https://max.ru/{cfg.USERNAME_MAX_BOT}"
            f"?start={p.get('participant_id')}"
        )

        buf = BytesIO(generate_qr_bytes(qr_link))
        buffers.append(buf)

        img = XLImage(buf)
        img.width = _QR_PX
        img.height = _QR_PX
        ws.add_image(img, f"D{idx}")

    out = BytesIO()
    wb.save(out)
    return out.getvalue()