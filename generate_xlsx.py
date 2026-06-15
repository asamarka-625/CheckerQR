from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Участники"

headers = ["ФИО", "Номер телефона", "Дополнительная информация"]
ws.append(headers)

# пример строки
ws.append(["Иванов Иван Иванович", "+7 (999) 123-45-67", "VIP-гость"])

# оформление заголовка
header_fill = PatternFill("solid", fgColor="1677FF")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 40

import os
os.makedirs("web_app/src/static/files", exist_ok=True)
wb.save("web_app/src/static/files/participants_template.xlsx")
print("ok")