#!/usr/bin/env python3
"""
Читает Excel-файл, где в одном из столбцов вставлены QR-коды (как картинки),
распознаёт зашитую в каждый QR ссылку и сохраняет новую таблицу,
в которой столбец с QR-кодами заменён на текстовую ссылку.

Запуск:
    python qr_to_links.py входной.xlsx [выходной.xlsx]

Если выходной файл не указан, рядом создаётся <имя>_links.xlsx
"""

import sys
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image


# --- распознавание QR ----------------------------------------------------------

def decode_qr(image_bytes):
    """Возвращает строку, зашитую в QR, или None. Сначала pyzbar, потом OpenCV."""
    img = Image.open(BytesIO(image_bytes))

    # 1) pyzbar (обычно точнее)
    try:
        from pyzbar.pyzbar import decode
        results = decode(img.convert("RGB"))
        if results:
            return results[0].data.decode("utf-8", errors="replace")
    except Exception:
        pass

    # 2) OpenCV как запасной вариант
    try:
        import cv2
        import numpy as np
        arr = cv2.cvtColor(np.array(img.convert("RGB")), cv2.COLOR_RGB2BGR)
        detector = cv2.QRCodeDetector()
        data, _, _ = detector.detectAndDecode(arr)
        if data:
            return data
    except Exception:
        pass

    return None


# --- извлечение картинок из листа ----------------------------------------------

def extract_uuid(text):
    """Из ссылки вида .../...?start=<uuid> возвращает только <uuid>.
    Если стандартный UUID найден в строке — возвращает его; иначе берёт
    значение параметра start; иначе — текст после последнего '='."""
    if not text:
        return text
    import re
    from urllib.parse import urlparse, parse_qs
    m = re.search(
        r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-"
        r"[0-9a-fA-F]{4}-[0-9a-fA-F]{12}", text)
    if m:
        return m.group(0)
    try:
        qs = parse_qs(urlparse(text).query)
        if "start" in qs and qs["start"]:
            return qs["start"][0]
    except Exception:
        pass
    return text.rsplit("=", 1)[-1] if "=" in text else text


def get_image_bytes(image):
    """Достаёт байты картинки из объекта openpyxl Image (разные версии)."""
    try:
        return image._data()
    except Exception:
        ref = image.ref
        if hasattr(ref, "read"):
            ref.seek(0)
            return ref.read()
        return Path(ref).read_bytes()


def image_anchor_row(image):
    """0-индексированный номер строки, к которой привязана картинка."""
    anchor = image.anchor
    frm = getattr(anchor, "_from", None)
    if frm is not None:
        return frm.row  # 0-based
    # анкор задан строкой вида "D2"
    if isinstance(anchor, str):
        from openpyxl.utils.cell import coordinate_to_tuple
        return coordinate_to_tuple(anchor)[0] - 1
    raise ValueError("Не удалось определить строку для картинки")


def build_row_to_url(ws):
    """{номер_строки(1-based): ссылка}"""
    mapping = {}
    for image in getattr(ws, "_images", []):
        try:
            row0 = image_anchor_row(image)
        except Exception:
            continue
        url = decode_qr(get_image_bytes(image))
        url = extract_uuid(url)
        # строка в Excel 1-based; если QR занимает несколько строк по высоте,
        # привязываемся к верхней — это и есть строка записи
        mapping[row0 + 1] = url
    return mapping


# --- основная логика -----------------------------------------------------------

def convert(in_path, out_path, qr_col_letter=None, link_header="UUID"):
    wb = load_workbook(in_path)
    ws = wb.active

    row_to_url = build_row_to_url(ws)
    if not row_to_url:
        print("ВНИМАНИЕ: на листе не найдено ни одной картинки QR-кода.")

    # определяем столбец с QR: либо указан вручную, либо тот, где сидят картинки
    if qr_col_letter:
        qr_col = ord(qr_col_letter.upper()) - ord("A") + 1
    else:
        cols = [img.anchor._from.col for img in getattr(ws, "_images", [])
                if getattr(img.anchor, "_from", None) is not None]
        qr_col = (min(cols) + 1) if cols else ws.max_column

    header_row = 1
    max_col = ws.max_column

    # читаем шапку и данные, пропуская столбец с QR-кодами
    out = Workbook()
    out_ws = out.active
    out_ws.title = ws.title

    # стили под исходное оформление
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_font = Font(name="Arial", bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="ED7D31")
    cell_font = Font(name="Arial")
    wrap = Alignment(vertical="center", wrap_text=True)

    def kept_columns():
        """Список исходных столбцов, которые переносим (без QR-столбца)."""
        return [c for c in range(1, max_col + 1) if c != qr_col]

    # шапка
    out_col = 1
    for c in kept_columns():
        val = ws.cell(row=header_row, column=c).value
        cell = out_ws.cell(row=header_row, column=out_col, value=val)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        out_col += 1
    # столбец со ссылкой (вместо QR)
    link_col_out = out_col
    cell = out_ws.cell(row=header_row, column=link_col_out, value=link_header)
    cell.font = head_font
    cell.fill = head_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    # данные
    for r in range(header_row + 1, ws.max_row + 1):
        out_col = 1
        row_has_data = False
        for c in kept_columns():
            val = ws.cell(row=r, column=c).value
            if val is not None:
                row_has_data = True
            cell = out_ws.cell(row=r, column=out_col, value=val)
            cell.font = cell_font
            cell.alignment = wrap
            cell.border = border
            out_col += 1
        url = row_to_url.get(r)
        cell = out_ws.cell(row=r, column=link_col_out, value=url)
        cell.font = cell_font
        cell.alignment = wrap
        cell.border = border
        if url:
            row_has_data = True
        # копируем высоту строки
        if r in ws.row_dimensions:
            out_ws.row_dimensions[r].height = ws.row_dimensions[r].height

    # ширины столбцов: переносим исходные, для столбца ссылки ставим пошире
    out_col = 1
    for c in kept_columns():
        letter_in = get_column_letter(c)
        letter_out = get_column_letter(out_col)
        if letter_in in ws.column_dimensions and ws.column_dimensions[letter_in].width:
            out_ws.column_dimensions[letter_out].width = ws.column_dimensions[letter_in].width
        out_col += 1
    out_ws.column_dimensions[get_column_letter(link_col_out)].width = 45

    out.save(out_path)
    return row_to_url


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    in_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else \
        in_path.with_name(in_path.stem + "_links.xlsx")

    mapping = convert(in_path, out_path)
    print(f"Готово: {out_path}")
    print(f"Распознано QR-кодов: {sum(1 for v in mapping.values() if v)} "
          f"из {len(mapping)}")
    for row, url in sorted(mapping.items()):
        print(f"  строка {row}: {url or '— не распознан —'}")


if __name__ == "__main__":
    main()