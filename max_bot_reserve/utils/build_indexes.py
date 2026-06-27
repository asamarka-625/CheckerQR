#!/usr/bin/env python3
"""Строит два индекса из таблицы с колонками ФИО, Доп. информация, Код, UUID."""

from openpyxl import load_workbook


def build_indexes(path, sheet=None, header_row=1):
    """
    Читает xlsx и возвращает кортеж из двух словарей:

      by_code: { код: {"фио": ..., "доп_информация": ..., "uuid": ...} }
      by_uuid: { uuid: {"фио": ..., "доп_информация": ..., "код": ...} }

    Колонки находятся по названиям в строке заголовка (header_row),
    поэтому порядок столбцов в файле не важен.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    import re

    def norm(s):
        return re.sub(r"[^0-9a-zа-я]+", "", str(s).lower().replace("ё", "е"))

    # карта "нормализованный заголовок -> номер столбца"
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val is not None and str(val).strip():
            headers[norm(val)] = col

    def col_of(*substrings):
        """Первый столбец, в нормализованном заголовке которого встречается
        любая из подстрок."""
        subs = [norm(s) for s in substrings]
        for key, col in headers.items():
            if any(sub in key for sub in subs):
                return col
        raise KeyError(
            f"Не найден столбец по {substrings}. "
            f"Доступные заголовки: {list(headers)}")

    c_fio = col_of("фио")
    c_info = col_of("информац")
    c_code = col_of("код")
    c_uuid = col_of("uuid", "ссылк")

    by_code, by_uuid = {}, {}
    for r in range(header_row + 1, ws.max_row + 1):
        fio = ws.cell(r, c_fio).value
        info = ws.cell(r, c_info).value
        code = ws.cell(r, c_code).value
        uid = ws.cell(r, c_uuid).value
        if fio is None and code is None and uid is None:
            continue  # пустая строка
        if code is not None:
            by_code[code] = {"фио": fio, "доп_информация": info, "uuid": uid}
        if uid is not None:
            by_uuid[uid] = {"фио": fio, "доп_информация": info, "код": code}

    return by_code, by_uuid